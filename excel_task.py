import openpyxl
import os
import re

header_file = "header.h"
#get the current working directory to check for all the header files in it
cwd = os.getcwd()
# list all the header files
headers_files= [files for files in os.listdir(cwd) if files.endswith('.h')]
#create excel workbook
ex_workbook= openpyxl.Workbook()
#create excel new worksheet
ex_worksheet= ex_workbook.create_sheet(title= "new sheet")
#or ex_worksheet= ex_workbook.active()

#search for the wanted header file
for header_file in headers_files:
    with open(header_file,'r') as file:
        header_file_data =file.read()
    # Find all function prototypes using findall
        function_prototypes = re.findall(r"(\w+ \w+\((.*?)\);)", header_file_data)
        #write function prototypes in sheet
        for i, function_prototype in enumerate(function_prototypes): #enumerate using to keep track of index
            ex_worksheet.cell(row=i+1,column=1).value= "IDX" + str(i)
            ex_worksheet.cell(row=i+1,column=2).value= function_prototype[0]

#save the excel sheet
ex_workbook.save("function_prototypes.xlsx")

#r stands for raw string (contains special characters no problem)
#\w+ for function type (ex- void)
#\w+\() for next word which is function name()
#(.*?) matches any character, including newlines
#\);) the function paranthesis

